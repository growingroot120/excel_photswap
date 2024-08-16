import os
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from io import BytesIO

# Paths to files
source_excel_1 = "MAP+Liaison+Report+7-20-24.xlsx"
source_excel_2 = "MAP+Liaison+Report+B+shift+7-21-24.xlsx"  # Path to the additional source Excel file
destination_excel = "06.07.24.xlsx"
image_folder = "ExtractImages"

# Ensure the output directory exists
if not os.path.exists(image_folder):
    os.makedirs(image_folder)

# Load the source workbooks
source_workbook_1 = load_workbook(source_excel_1)
source_workbook_2 = load_workbook(source_excel_2)

# Load or create the destination workbook
if os.path.exists(destination_excel):
    destination_workbook = load_workbook(destination_excel)
else:
    destination_workbook = Workbook()

# Ensure the destination workbook has sheets from 1 to 60
for sheet_number in range(1, 61):
    sheet_name = str(sheet_number)
    if sheet_name not in destination_workbook.sheetnames:
        destination_workbook.create_sheet(title=sheet_name)

# Remove the default sheet if it exists (usually named 'Sheet' in a new workbook)
if 'Sheet' in destination_workbook.sheetnames:
    del destination_workbook['Sheet']

def extract_images(source_workbook, folder_name, start_sheet_num, end_sheet_num):
    for sheet_number in range(start_sheet_num, end_sheet_num + 1):
        sheet_name = f"PHOTO {sheet_number}"
        if sheet_name in source_workbook.sheetnames:
            sheet = source_workbook[sheet_name]
            for index, drawing in enumerate(sheet._images):
                img = PILImage.open(BytesIO(drawing._data()))
                img = img.resize((320, 365), PILImage.LANCZOS)
                image_path = os.path.join(folder_name, f"{sheet_name}-Image-{index}.png")
                img.save(image_path)

# Extract images from the first source workbook
extract_images(source_workbook_1, image_folder, 1, 30)

# Determine the last sheet number used by the first source workbook
last_used_sheet_number = 0
for sheet_number in range(1, 31):
    sheet_name = f"PHOTO {sheet_number}"
    if sheet_name in source_workbook_1.sheetnames and source_workbook_1[sheet_name]._images:
        last_used_sheet_number = sheet_number

# Extract images from the second source workbook, starting from the next available sheet number
extract_images(source_workbook_2, image_folder, last_used_sheet_number + 1, last_used_sheet_number + 30)

print("Images extracted successfully from both source workbooks.")

# Insert images into destination workbook
for sheet_number in range(1, 61):
    sheet_name = str(sheet_number)
    sheet = destination_workbook[sheet_name]

    # Clear existing images
    for img in sheet._images:
        sheet._images.remove(img)

    for image_file in os.listdir(image_folder):
        if image_file.startswith(f"PHOTO {sheet_number}-Image-"):
            img_path = os.path.join(image_folder, image_file)
            img = OpenpyxlImage(img_path)
            img.width, img.height = 320, 365  # Set image size in pixels
            sheet.add_image(img, 'G9')  # Adjust the position as needed

# Save the updated destination workbook
destination_workbook.save(destination_excel)

# Remove the ExtractImages folder
shutil.rmtree(image_folder)

print("Images copied to destination workbook successfully and extracted images folder removed.")
